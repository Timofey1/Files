import xlrd
from openpyxl import load_workbook
from datetime import datetime


def convert(month_number):
    month = {'01': "январь", '02': "февраль", '03': "март", '04': "апрель", '05': "май", '06': "июнь", '07': "июль",
             '08': "август", '09': "сентябрь", '10': "октябрь", '11': "ноябрь", '12': "декабрь"}
    return month[month_number]

d = str(datetime.now())
d = d[:d.find(" ")]
d = d.split("-")

month = convert(d[1])
date = int(d[2])
mas = ['Бешкуров Тимофей Борисович']

for fio in mas:
    workbook = xlrd.open_workbook('export.xlsx')
    sheet = workbook.sheet_by_index(0)
    curcol = 1    
    for i in range(sheet.nrows):
        data = sheet.cell_value(i, 0)
        if data == fio:
            currow = i
            break
    for i in range(sheet.ncols):
        data = sheet.cell_value(0, i)
        if data == month:
            curcol = i
            break
        
    curcol += date
    
    book=load_workbook('export.xlsx')
    sheet = book.active
    currow += 1
    sheet.cell(row=currow, column=curcol).value = "X"
    book.save('export.xlsx')

    bookmain = load_workbook('10kl.xlsx')
    sheetmain = bookmain.active
    sheetmain.cell(row=currow, column=curcol).value = "H"
    bookmain.save("10kl.xlsx")
    
