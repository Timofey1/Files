import xlrd
from openpyxl import load_workbook
from datetime import datetime

def convert(month_number):
    month = {'01': "январь", '02': "февраль", '03': "март", '04': "апрель", '05': "май", '06': "июнь", '07': "июль",
             '08': "август", '09': "сентябрь", '10': "октябрь", '11': "ноябрь", '12': "декабрь"}
    return month[month_number]

d = str(datetime.now())
d = d[:d.find(" ")]
d = d.split("-")
month = convert(d[1])
date = int(d[2])

workbook = xlrd.open_workbook('10kl.xlsx')
sheet = workbook.sheet_by_index(0)

for i in range(sheet.ncols):
    data = sheet.cell_value(0, i)
    if data == month:
        break

curcol = i + date

for i in range(3, sheet.nrows + 1):
    book = load_workbook('export.xlsx')
    sheet = book.active
    sheet.cell(row=i, column=curcol).value = "H"
    book.save('export.xlsx')
